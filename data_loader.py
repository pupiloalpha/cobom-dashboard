"""Leitura e processamento inicial dos arquivos do dashboard."""

import io
from typing import Any

import chardet
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from utils.helpers import COLUMN_MAPPING, normalize_column_names, parse_coordinate, parse_datetime_series

XLSX_COLUMNS = [
    "chamada_numero", "reds", "data_hora_criacao", "hora_criacao",
    "Chamada_atendimentos.local_do_fato", "Chamada_atendimentos.local_latitude",
    "Chamada_atendimentos.local_longitude", "Chamada_atendimentos.natureza_codigo",
    "Chamada_atendimentos.natureza_descricao", "Chamada_atendimentos.unidade_servico_codigo",
    "Chamada_atendimentos.unidade_servico_nome", "Empenhos.recurso_codigo_prefixo",
    "Chamada_atendimentos.chamada_classificacao_descricao", "data_classificacao",
    "hora_classificacao", "estado_chamada", "Chamada_atendimentos.local_municipio_id",
    "Chamada_atendimentos.local_municipio_nome",
]

CSV_COLUMNS = [
    "chamada_numero", "reds", "data_hora_criacao", "Chamada_atendimentos.local_do_fato",
    "Chamada_atendimentos.local_latitude", "Chamada_atendimentos.local_longitude",
    "Chamada_atendimentos.natureza_descricao", "Chamada_atendimentos.unidade_servico_nome",
    "Empenhos.recurso_codigo_prefixo", "alerta", "destaque", "envolve_autoridade",
    "Chamada_atendimentos.chamada_classificacao_descricao", "situacao",
    "data_hora_situacao_atual", "evento_associado",
]


def read_uploaded_file(uploaded_file: Any) -> pd.DataFrame:
    """Le CSV ou XLSX usando o conteudo do upload, sem cachear o objeto recebido."""
    raw = uploaded_file.getvalue()
    filename = uploaded_file.name.lower().strip()
    is_excel = filename.endswith((".xlsx", ".xlsm", ".xslx")) or raw[:4] == b"PK\x03\x04"
    if is_excel:
        return _normalize_excel_schema(_read_excel_with_openpyxl(raw))

    detected = chardet.detect(raw[:100_000]).get("encoding") or "utf-8"
    attempts = list(dict.fromkeys([detected, "utf-8-sig", "utf-8", "cp1252", "latin-1"]))
    last_error = None
    for encoding in attempts:
        try:
            df = pd.read_csv(
                io.BytesIO(raw), sep=None, engine="python", encoding=encoding,
                dtype=str, on_bad_lines="skip",
            )
            return _normalize_fixed_schema(normalize_column_names(df), CSV_COLUMNS)
        except (UnicodeDecodeError, pd.errors.ParserError, ValueError) as error:
            last_error = error
    raise ValueError(f"CSV nao pode ser lido: {last_error}")


def _read_excel_with_openpyxl(raw: bytes) -> pd.DataFrame:
    """Le a aba COBOM e identifica o cabecalho pelos nomes conhecidos."""
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        worksheet = next(
            (sheet for sheet in workbook.worksheets if sheet.title.strip().lower() == "bd_cobom"),
            None,
        )
        if worksheet is None:
            worksheet = next(
                (sheet for sheet in workbook.worksheets if sheet.max_row and sheet.max_column),
                None,
            )
        if worksheet is None:
            return pd.DataFrame()

        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        known_headers = set(COLUMN_MAPPING) | set(XLSX_COLUMNS) | {
            "Reds.reds_numero",
            "chamada_data_inclusao",
            "chamada_hora_inclusao",
        }
        header_index = max(
            range(len(rows)),
            key=lambda index: sum(
                str(value).strip() in known_headers
                for value in rows[index]
                if value is not None
            ),
            default=0,
        )
        header_score = sum(
            str(value).strip() in known_headers
            for value in rows[header_index]
            if value is not None
        )
        width = max((len(row) for row in rows), default=0)
        rows = [row + [None] * (width - len(row)) for row in rows]

        if header_score:
            header = [
                str(value).strip() if value is not None and str(value).strip() else f"coluna_{index + 1}"
                for index, value in enumerate(rows[header_index])
            ]
            return pd.DataFrame(rows[header_index + 1:], columns=header)

        # Fallback for exports whose header names are absent or changed.
        candidates = []
        for worksheet in workbook.worksheets:
            if not worksheet.max_row or not worksheet.max_column:
                continue
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            first_data_index = next(
                (index for index, row in enumerate(rows) if any(value is not None for value in row)),
                None,
            )
            if first_data_index is not None:
                header_index = max(
                    range(len(rows)),
                    key=lambda index: (
                        sum(value is not None for value in rows[index]),
                        -index,
                    ),
                    default=first_data_index,
                )
                header_score = sum(value is not None for value in rows[header_index])
                candidates.append((header_score, worksheet, rows, header_index))

        if not candidates:
            return pd.DataFrame()

        _, _, rows, header_index = max(candidates, key=lambda item: item[0])
        width = max((len(row) for row in rows), default=0)
        rows = [row + [None] * (width - len(row)) for row in rows]
        data = rows[header_index + 1:]
        return pd.DataFrame(data).reindex(columns=range(len(XLSX_COLUMNS))).set_axis(XLSX_COLUMNS, axis=1)
    except Exception as error:
        raise ValueError(f"Excel nao pode ser lido: {error}") from error


def _normalize_excel_schema(df: pd.DataFrame) -> pd.DataFrame:
    """Preserva cabecalhos XLSX conhecidos e usa posicoes somente no fallback."""
    normalized = normalize_column_names(df)
    known_columns = set(XLSX_COLUMNS) | {
        "chamada_data_inclusao",
        "chamada_hora_inclusao",
        "Chamada_atendimentos.chamada_classificacao_data",
        "Chamada_atendimentos.chamada_classificacao_hora",
    }
    if any(column in normalized.columns for column in known_columns):
        return _normalize_fixed_schema_by_name(normalized)
    return _normalize_fixed_schema(normalized, XLSX_COLUMNS)


def _normalize_fixed_schema_by_name(df: pd.DataFrame) -> pd.DataFrame:
    """Converte aliases do Excel sem descartar colunas reconhecidas pelo cabecalho."""
    aliases = {
        "Reds.reds_numero": "reds",
        "Chamada_atendimentos.chamada_classificacao_data": "data_classificacao",
        "Chamada_atendimentos.chamada_classificacao_hora": "hora_classificacao",
    }
    result = df.rename(columns={source: target for source, target in aliases.items()})
    if "chamada_data_inclusao" in result.columns:
        result = result.rename(columns={
            "chamada_data_inclusao": "data_hora_criacao",
            "chamada_hora_inclusao": "hora_criacao",
        })
    if "data_hora_criacao" not in result.columns and "chamada_data_inclusao" in df.columns:
        result["data_hora_criacao"] = df["chamada_data_inclusao"]
    if {"data_hora_criacao", "hora_criacao"}.issubset(result.columns):
        result["data_hora_criacao"] = (
            result["data_hora_criacao"].astype("string").str.strip()
            + " "
            + result["hora_criacao"].astype("string").str.strip()
        )
    if "data_hora_situacao_atual" not in result.columns and {
        "data_classificacao", "hora_classificacao"
    }.issubset(result.columns):
        result["data_hora_situacao_atual"] = (
            result["data_classificacao"].astype("string").str.strip()
            + " "
            + result["hora_classificacao"].astype("string").str.strip()
        )
    return result


def _normalize_fixed_schema(df: pd.DataFrame, fixed_columns: list[str]) -> pd.DataFrame:
    """Converte os campos especificos dos esquemas XLSX e CSV para o modelo comum."""
    result = df.iloc[:, :len(fixed_columns)].copy()
    if result.shape[1] < len(fixed_columns):
        for index in range(result.shape[1], len(fixed_columns)):
            result[index] = pd.NA
    result.columns = fixed_columns
    aliases = {
        "Reds.reds_numero": "reds",
        "chamada_data_inclusao": "data_hora_criacao",
        "chamada_hora_inclusao": "hora_criacao",
        "Chamada_atendimentos.chamada_classificacao_data": "data_classificacao",
        "Chamada_atendimentos.chamada_classificacao_hora": "hora_classificacao",
    }
    result = result.rename(columns={source: target for source, target in aliases.items() if source in result})

    if "data_hora_criacao" in result.columns and "hora_criacao" in result.columns:
        result["data_hora_criacao"] = (
            result["data_hora_criacao"].astype("string").str.strip()
            + " "
            + result["hora_criacao"].astype("string").str.strip()
        )
    if "data_classificacao" in result.columns and "hora_classificacao" in result.columns:
        result["data_hora_situacao_atual"] = (
            result["data_classificacao"].astype("string").str.strip()
            + " "
            + result["hora_classificacao"].astype("string").str.strip()
        )

    return result


def _numeric_coordinates(series: pd.Series, max_abs: float) -> pd.Series:
    text = series.astype("string").str.strip()
    numeric = pd.to_numeric(text.str.replace(",", ".", regex=False), errors="coerce").astype("float64")

    # Alguns exports removem o separador decimal e variam a quantidade de casas.
    # Testa escalas decimais ate encontrar a primeira coordenada plausivel.
    integer_like = numeric.notna() & numeric.mod(1).eq(0)
    unformatted = numeric.abs().gt(max_abs) & (
        ~text.str.contains(r"[.,]", regex=True, na=False) | integer_like
    )
    for index in numeric.index[unformatted]:
        value = numeric.at[index]
        for scale in range(1, 10):
            candidate = value / 10 ** scale
            if abs(candidate) <= max_abs:
                numeric.at[index] = candidate
                break

    fallback = numeric.isna() & series.notna()
    if fallback.any():
        numeric.loc[fallback] = series.loc[fallback].map(
            lambda value: parse_coordinate(value, max_abs)
        )
    return numeric.where(numeric.abs().le(max_abs))


def process_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Cria campos derivados usados por filtros, metricas e graficos."""
    result = df.copy()
    if "data_hora_criacao" in result.columns:
        created = parse_datetime_series(result["data_hora_criacao"])
        result["chamada_data_inclusao"] = created.dt.normalize()
        result["chamada_hora_inclusao"] = pd.to_timedelta(created.dt.time.astype(str), errors="coerce")
        result["data_hora"] = created

    for column, max_abs in (
        ("Chamada_atendimentos.local_latitude", 90),
        ("Chamada_atendimentos.local_longitude", 180),
    ):
        if column in result.columns:
            result[column] = _numeric_coordinates(result[column], max_abs)

    local_column = "Chamada_atendimentos.local_do_fato"
    if local_column in result.columns:
        result["Chamada_atendimentos.local_municipio_nome"] = result[local_column].map(
            lambda value: value if pd.isna(value) else str(value).split(" - ")[-1].strip()
        )

    if "chamada_data_inclusao" in result.columns:
        result = result.dropna(subset=["chamada_data_inclusao"])
        result["ano"] = result["chamada_data_inclusao"].dt.year
        result["mes"] = result["chamada_data_inclusao"].dt.month
        result["mes_ano"] = result["chamada_data_inclusao"].dt.to_period("M").astype(str)
        result["hora"] = (result["chamada_hora_inclusao"].dt.total_seconds() // 3600).astype("Int64")
        result["dia_semana"] = result["chamada_data_inclusao"].dt.dayofweek

    if "data_hora_situacao_atual" in result.columns:
        result["data_hora_fim"] = parse_datetime_series(result["data_hora_situacao_atual"])
    else:
        result["data_hora_fim"] = pd.NaT
    return result


def load_uploaded_data(uploaded_file: Any) -> pd.DataFrame:
    return process_dataframe(read_uploaded_file(uploaded_file))


def _filters_key(filters: dict[str, Any]) -> tuple:
    return tuple((key, tuple(value) if isinstance(value, list) else value) for key, value in sorted(filters.items()))


@st.cache_data(show_spinner=False)
def _apply_filters_cached(df: pd.DataFrame, filters_key: tuple) -> pd.DataFrame:
    result = df
    for column, selected in filters_key:
        if selected and column in result.columns:
            if column == "Empenhos.recurso_codigo_prefixo":
                resources = result[column].fillna("").astype(str).str.replace(" / ", ",", regex=False).str.split(",").explode()
                matching = resources.str.strip().isin(selected).groupby(level=0).any()
                result = result.loc[result.index.intersection(matching[matching].index)]
            else:
                result = result[result[column].isin(selected)]
    return result


def apply_filters(df: pd.DataFrame, filters_dict: dict[str, Any]) -> pd.DataFrame:
    """Aplica filtros hashable, mantendo o DataFrame original intacto."""
    return _apply_filters_cached(df, _filters_key(filters_dict)).copy()
